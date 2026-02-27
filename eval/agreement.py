#!/usr/bin/env python3

import openpyxl
from sklearn.metrics import cohen_kappa_score


def load_scores(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    all_scores = {}          # {team: {qid:0/1}}
    for sheetname in wb.sheetnames:
        if not sheetname.startswith("Team"):
            continue
        sheet = wb[sheetname]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            qid = row[0]
            if qid is None:          # empty row
                continue
            # try the "Annotator Score" column first (index 5), else use Pass/Fail
            score = row[5]
            if score is None:
                if row[3] == "Pass (1)":
                    score = 1
                elif row[4] == "Fail (0)":
                    score = 0
            if score is None:
                continue   # not evaluated
            all_scores.setdefault(sheetname, {})[qid] = int(score)
    return all_scores


def scores_per_project(annotator_name, scores_dict):
    """Per ogni progetto (team): punteggio medio (media degli 0/1 sulle categorie)."""
    out = []
    for team, q_scores in sorted(scores_dict.items()):
        vals = list(q_scores.values())
        mean = sum(vals) / len(vals) if vals else 0.0
        out.append((team, mean, len(vals)))
    return out


def pct_score_one_per_category(scores_dict):
    """Per ogni categoria (qid): percentuale di progetti con score 1."""
    by_qid = {}
    for team, q_scores in scores_dict.items():
        for qid, score in q_scores.items():
            by_qid.setdefault(qid, []).append(score)
    return {
        qid: (sum(s for s in vals if s == 1) / len(vals) * 100.0 if vals else 0.0, len(vals))
        for qid, vals in by_qid.items()
    }


def main():
    m1 = load_scores("eval/matteo.xlsx")
    m2 = load_scores("eval/nicco.xlsx")

    # build list of paired scores
    common = []
    for team, scores in m1.items():
        if team in m2:
            for qid, s1 in scores.items():
                s2 = m2[team].get(qid)
                if s2 is not None:
                    common.append((s1, s2))

    if not common:
        print("Nessun punteggio in comune; accertati di aver compilato i fogli.")
        return

    r1, r2 = zip(*common)
    kappa_all = cohen_kappa_score(r1, r2)
    print(f"kappa totale (tutti i quesiti): {kappa_all:.3f}")

    # per quesito
    by_q = {}
    for team, scores in m1.items():
        if team in m2:
            for qid, s1 in scores.items():
                if qid in m2[team]:
                    by_q.setdefault(qid, []).append((s1, m2[team][qid]))
    print("kappa per quesito:")
    for qid, pairs in sorted(by_q.items()):
        v1, v2 = zip(*pairs)
        print(f"  {qid}: {cohen_kappa_score(v1, v2):.3f}")

    # --- Punteggi per singolo progetto ---
    print("\n" + "=" * 60)
    print("PUNTEGGIO PER SINGOLO PROGETTO (media sulle categorie)")
    print("=" * 60)
    for name, data in [("matteo", m1), ("nicco", m2)]:
        print(f"\n--- {name} ---")
        rows = scores_per_project(name, data)
        for team, mean, n in rows:
            print(f"  {team}: {mean:.3f}  (n={n} categorie)")
        if rows:
            tot_mean = sum(m for _, m, _ in rows) / len(rows)
            print(f"  Media tra progetti ({name}): {tot_mean:.3f}")

    # --- Punteggio medio totale (tutti i progetti × tutte le categorie) ---
    print("\n" + "=" * 60)
    print("PUNTEGGIO MEDIO TOTALE (tutti i progetti e categorie)")
    print("=" * 60)
    for name, data in [("matteo", m1), ("nicco", m2)]:
        all_vals = [s for team_scores in data.values() for s in team_scores.values()]
        if all_vals:
            total_mean = sum(all_vals) / len(all_vals)
            print(f"  {name}: {total_mean:.3f}  (su {len(all_vals)} valutazioni)")
    # media congiunta (solo celle in comune)
    r1_flat, r2_flat = list(r1), list(r2)
    if r1_flat:
        print(f"  media matteo (solo celle in comune): {sum(r1_flat)/len(r1_flat):.3f}")
    if r2_flat:
        print(f"  media nicco (solo celle in comune): {sum(r2_flat)/len(r2_flat):.3f}")

    # --- Percentuale score 1 per categoria (per annotatore) ---
    print("\n" + "=" * 60)
    print("PERCENTUALE DI SCORE 1 PER CATEGORIA (per annotatore)")
    print("=" * 60)
    for name, data in [("matteo", m1), ("nicco", m2)]:
        pct = pct_score_one_per_category(data)
        print(f"\n--- {name} ---")
        for qid in sorted(pct.keys()):
            p, n = pct[qid]
            print(f"  {qid}: {p:.1f}%  (n={n} progetti)")
        if pct:
            count_ones = sum(s for team_scores in data.values() for s in team_scores.values() if s == 1)
            total_n = sum(len(v) for v in data.values())
            print(f"  TOTALE (% valutazioni con score 1): {count_ones/total_n*100:.1f}%" if total_n else "")


if __name__ == "__main__":
    main()
